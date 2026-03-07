"""
BPExtractor: Extracts clinical code mapping data from the Best Practice
reference data Excel file.

Each sheet in the Excel file represents a FHIR entity type and contains
BP (Best Practice) source codes alongside their mapped terminology codes
(SNOMED CT AU, ICD-10-AM, AMT, LOINC, AIR, DICOM).

The sheets have complex, non-uniform layouts -- headers are not always in
row 1, multiple data sections may sit side-by-side, and annotation rows
are interspersed.  This module handles those variations per-sheet.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Sheet-specific layout definitions
# ---------------------------------------------------------------------------
# Each entry describes one *section* of a sheet.
#   header_row : 1-based row number that contains column headers
#   data_start : 1-based row number where data begins
#   columns    : list of (column_letter, header_name) pairs
#   id_field   : which header_name holds the source code / ID
#   display_field : which header_name holds the human-readable display name
#   section_label : optional label to disambiguate when a sheet has multiple
#                   side-by-side sections

_SHEET_CONFIGS: dict[str, list[dict[str, Any]]] = {
    "AllergyIntolerance": [
        {
            "section_label": "Type1_Product",
            "header_row": 5,
            "data_start": 6,
            "columns": [("B", "PRODUCTNAMEID"), ("C", "PRODUCTNAME"), ("D", "SEARCHNAME")],
            "id_field": "PRODUCTNAMEID",
            "display_field": "PRODUCTNAME",
        },
        {
            "section_label": "Type2_Ingredient",
            "header_row": 5,
            "data_start": 6,
            "columns": [("F", "INGREDIENTID"), ("G", "INGREDIENTNAME"), ("H", "GENCODE")],
            "id_field": "INGREDIENTID",
            "display_field": "INGREDIENTNAME",
        },
        {
            "section_label": "Type3_DrugClass",
            "header_row": 5,
            "data_start": 6,
            "columns": [("J", "DRUGCLASSID"), ("K", "DESCRIPTION")],
            "id_field": "DRUGCLASSID",
            "display_field": "DESCRIPTION",
        },
        {
            "section_label": "Type4_Agent",
            "header_row": 5,
            "data_start": 6,
            "columns": [("N", "AGENTID"), ("O", "AGENTNAME")],
            "id_field": "AGENTID",
            "display_field": "AGENTNAME",
        },
        {
            "section_label": "Reactions",
            "header_row": 4,
            "data_start": 5,
            "columns": [("R", "REACTIONCODE"), ("S", "SnomedCode"), ("T", "Display")],
            "id_field": "REACTIONCODE",
            "display_field": "Display",
        },
    ],
    "Medication": [
        {
            "section_label": "FormCode",
            "header_row": 4,
            "data_start": 5,
            "columns": [
                ("A", "FORMCODE"),
                ("B", "COMPANYFORM"),
                ("C", "Code"),
                ("D", "Display"),
                ("E", "System"),
            ],
            "id_field": "FORMCODE",
            "display_field": "COMPANYFORM",
        },
        {
            "section_label": "Product",
            "header_row": 3,
            "data_start": 4,
            "columns": [
                ("G", "productID"),
                ("H", "GENERICNAME"),
                ("I", "PRODUCTNAMEID"),
                ("J", "PRODUCTNAME"),
                ("K", "SEARCHNAME"),
                ("L", "AMTMPPCODE"),
                ("M", "AMTcode"),
                ("N", "AMTpreferredTerm"),
                ("O", "DRUGNAME"),
            ],
            "id_field": "PRODUCTNAMEID",
            "display_field": "PRODUCTNAME",
        },
    ],
    "Immunization": [
        {
            "section_label": "Routes",
            "header_row": 5,
            "data_start": 6,
            "columns": [
                ("A", "ROUTECODE"),
                ("B", "System"),
                ("C", "CODE"),
                ("D", "Display"),
            ],
            "id_field": "ROUTECODE",
            "display_field": "Display",
        },
        {
            "section_label": "Vaccines_FHIR",
            "header_row": 5,
            "data_start": 6,
            "columns": [
                ("F", "VACCINEID"),
                ("G", "VACCINENAME"),
                ("H", "ACIRCODE"),
                ("I", "FHIR_CODE"),
                ("J", "FHIR_Display"),
                ("K", "System"),
            ],
            "id_field": "VACCINEID",
            "display_field": "VACCINENAME",
        },
        {
            "section_label": "Vaccines_AIR",
            "header_row": 5,
            "data_start": 6,
            "columns": [
                ("M", "VACCINEID"),
                ("N", "VACCINENAME"),
                ("O", "ACIRCODE"),
                ("P", "AIR_CODE"),
                ("Q", "AIR_Display"),
            ],
            "id_field": "VACCINEID",
            "display_field": "VACCINENAME",
        },
        {
            "section_label": "Sites",
            "header_row": 5,
            "data_start": 6,
            "columns": [
                ("S", "SITECODE"),
                ("T", "SITE"),
                ("U", "Code"),
                ("V", "Display"),
            ],
            "id_field": "SITECODE",
            "display_field": "SITE",
        },
    ],
    "Procedures": [
        {
            "section_label": None,
            "header_row": 1,
            "data_start": 2,
            "columns": [("A", "TERMNAMEID"), ("B", "TERMID"), ("C", "TERMNAME")],
            "id_field": "TERMNAMEID",
            "display_field": "TERMNAME",
        },
    ],
    "Conditions": [
        {
            "section_label": None,
            "header_row": 1,
            "data_start": 2,
            "columns": [("A", "TERMNAMEID"), ("B", "TERMID"), ("C", "TERMNAME")],
            "id_field": "TERMNAMEID",
            "display_field": "TERMNAME",
        },
    ],
    "Patient": [
        {
            "section_label": "Gender",
            "header_row": 3,
            "data_start": 4,
            "columns": [
                ("A", "Value"),
                ("B", "Meaning"),
                ("C", "System"),
                ("D", "Display"),
                ("E", "code"),
            ],
            "id_field": "Value",
            "display_field": "Meaning",
            "data_end": 10,
        },
        {
            "section_label": "IndigenousStatus",
            "header_row": 3,
            "data_start": 4,
            "columns": [
                ("H", "ETHNICCODE"),
                ("I", "ETHNICTYPE"),
                ("J", "System"),
                ("K", "Display"),
                ("L", "code"),
            ],
            "id_field": "ETHNICCODE",
            "display_field": "ETHNICTYPE",
            "data_end": 8,
        },
        {
            "section_label": "ContactRelationship",
            "header_row": 13,
            "data_start": 14,
            "columns": [
                ("A", "Code"),
                ("B", "Display"),
                ("C", "System"),
            ],
            "id_field": "Code",
            "display_field": "Display",
        },
    ],
    "Person": [
        {
            "section_label": "Gender",
            "header_row": 3,
            "data_start": 4,
            "columns": [
                ("A", "Value"),
                ("B", "Meaning"),
                ("C", "System"),
                ("D", "Display"),
                ("E", "code"),
            ],
            "id_field": "Value",
            "display_field": "Meaning",
        },
    ],
    "PractitionerRole": [
        {
            "section_label": "Role",
            "header_row": 6,
            "data_start": 7,
            "columns": [
                ("A", "GROUPCODE"),
                ("B", "GROUPNAME"),
                ("C", "Code"),
                ("D", "Display"),
                ("E", "System"),
            ],
            "id_field": "GROUPCODE",
            "display_field": "GROUPNAME",
        },
    ],
    "Slot": [
        {
            "section_label": "Role",
            "header_row": 2,
            "data_start": 3,
            "columns": [
                ("A", "GROUPCODE"),
                ("B", "GROUPNAME"),
                ("C", "Code"),
                ("D", "Display"),
                ("E", "System"),
            ],
            "id_field": "GROUPCODE",
            "display_field": "GROUPNAME",
        },
    ],
    "DocumentReference": [
        {
            "section_label": "Type_Source1",
            "header_row": 2,
            "data_start": 3,
            "columns": [
                ("A", "BP"),
                ("B", "Code"),
                ("C", "Display"),
            ],
            "id_field": "BP",
            "display_field": "Display",
        },
        {
            "section_label": "Category_Source1",
            "header_row": 2,
            "data_start": 3,
            "columns": [
                ("E", "BP"),
                ("F", "Code"),
                ("G", "Display"),
            ],
            "id_field": "BP",
            "display_field": "Display",
        },
        {
            "section_label": "CorrespondenceOut",
            "header_row": 2,
            "data_start": 3,
            "columns": [
                ("I", "BP_template_subject"),
                ("J", "LOINC_Code_Type"),
                ("K", "Document_Type"),
                ("L", "LOINC_Code_Category"),
                ("M", "Document_Category"),
            ],
            "id_field": "BP_template_subject",
            "display_field": "Document_Type",
        },
        {
            "section_label": "SecurityLabel",
            "header_row": 2,
            "data_start": 3,
            "columns": [
                ("O", "BP_Code"),
                ("P", "Code"),
                ("Q", "System"),
                ("R", "Display"),
            ],
            "id_field": "BP_Code",
            "display_field": "Display",
        },
    ],
    "Encounter": [
        {
            "section_label": None,
            "header_row": 3,
            "data_start": 4,
            "columns": [
                ("A", "VISITCODE"),
                ("B", "VISITTYPE"),
                ("C", "Code"),
                ("D", "System"),
                ("E", "Display"),
            ],
            "id_field": "VISITCODE",
            "display_field": "VISITTYPE",
        },
    ],
    "Observation-PathologyResult": [
        {
            "section_label": "ResultName_BPCode",
            "header_row": 3,
            "data_start": 4,
            "columns": [
                ("A", "ResultName"),
                ("B", "BP_CODE"),
            ],
            "id_field": "BP_CODE",
            "display_field": "ResultName",
        },
        {
            "section_label": "LOINC_BPCode",
            "header_row": 3,
            "data_start": 4,
            "columns": [
                ("D", "RECORDID"),
                ("E", "RECORDSTATUS"),
                ("F", "LOINC"),
                ("G", "BPCODE"),
            ],
            "id_field": "BPCODE",
            "display_field": "LOINC",
        },
        {
            "section_label": "Comparator",
            "header_row": 3,
            "data_start": 4,
            "columns": [
                ("I", "Code"),
                ("J", "Display"),
                ("K", "Definition"),
            ],
            "id_field": "Code",
            "display_field": "Display",
        },
    ],
    "Observation": [
        {
            "section_label": "ObservationCodes",
            "header_row": 2,
            "data_start": 3,
            "columns": [
                ("A", "DATACODE"),
                ("B", "MODIFIER"),
                ("C", "CODE"),
                ("D", "DISPLAY"),
            ],
            "id_field": "DATACODE",
            "display_field": "DISPLAY",
        },
        {
            "section_label": "FHIRValues",
            "header_row": 3,
            "data_start": 4,
            "columns": [
                ("G", "Code"),
                ("H", "Display"),
            ],
            "id_field": "Code",
            "display_field": "Display",
        },
    ],
    "ImagingStudy": [
        {
            "section_label": "Modality",
            "header_row": 2,
            "data_start": 3,
            "columns": [
                ("A", "BP_Code"),
                ("B", "BP_Display"),
                ("C", "FHIR_Code"),
                ("D", "FHIR_Display"),
            ],
            "id_field": "BP_Code",
            "display_field": "BP_Display",
        },
        {
            "section_label": "DICOM_Modality",
            "header_row": 2,
            "data_start": 3,
            "columns": [
                ("G", "Code_Value"),
                ("H", "Code_Meaning"),
                ("I", "Display"),
            ],
            "id_field": "Code_Value",
            "display_field": "Display",
        },
    ],
    "Medication Request": [
        {
            "section_label": "DosageRoutes",
            "header_row": 2,
            "data_start": 3,
            "columns": [
                ("A", "ROUTEID"),
                ("B", "code"),
                ("C", "Display"),
                ("D", "System"),
            ],
            "id_field": "ROUTEID",
            "display_field": "Display",
        },
    ],
    "ServiceRequest": [
        {
            "section_label": None,
            "header_row": 2,
            "data_start": 3,
            "columns": [
                ("A", "BP_Code"),
                ("B", "FHIR_Code"),
                ("C", "FHIR_Display"),
                ("D", "FHIR_System"),
            ],
            "id_field": "BP_Code",
            "display_field": "FHIR_Display",
        },
    ],
}


def _col_letter_to_index(letter: str) -> int:
    """Convert a column letter (A, B, ..., Z, AA, ...) to a 1-based index."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


class BPExtractor:
    """Extracts Best Practice reference data from an Excel workbook.

    Parameters
    ----------
    file_path : str | Path
        Path to the source Excel file.
    """

    def __init__(self, file_path: str | Path) -> None:
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Source file not found: {self.file_path}")
        self._wb: openpyxl.Workbook | None = None

    # ------------------------------------------------------------------
    # Workbook lifecycle helpers
    # ------------------------------------------------------------------

    def _open(self) -> openpyxl.Workbook:
        if self._wb is None:
            self._wb = openpyxl.load_workbook(
                str(self.file_path), data_only=True
            )
        return self._wb

    def close(self) -> None:
        if self._wb is not None:
            self._wb.close()
            self._wb = None

    # context-manager support
    def __enter__(self) -> "BPExtractor":
        self._open()
        return self

    def __exit__(self, *exc: object) -> None:
        self.close()

    # ------------------------------------------------------------------
    # Internal: read a single section from a worksheet
    # ------------------------------------------------------------------

    @staticmethod
    def _clean_value(value: Any) -> Any:
        """Strip whitespace from string values and normalise blanks."""
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            # treat non-breaking spaces and empty strings as None
            if value in ("", "\xa0"):
                return None
            return value
        return value

    def _read_section(
        self,
        ws: Any,
        section_cfg: dict[str, Any],
        sheet_name: str,
    ) -> list[dict[str, Any]]:
        """Read one side-by-side section of a worksheet.

        Returns a list of normalised record dicts.
        """
        columns = section_cfg["columns"]
        id_field = section_cfg["id_field"]
        display_field = section_cfg["display_field"]
        data_start = section_cfg["data_start"]
        data_end = section_cfg.get("data_end", ws.max_row)
        section_label = section_cfg.get("section_label")

        records: list[dict[str, Any]] = []

        for row_idx in range(data_start, data_end + 1):
            raw_data: dict[str, Any] = {}
            all_none = True
            for col_letter, header_name in columns:
                col_idx = _col_letter_to_index(col_letter)
                cell_value = self._clean_value(ws.cell(row=row_idx, column=col_idx).value)
                raw_data[header_name] = cell_value
                if cell_value is not None:
                    all_none = False

            if all_none:
                continue

            source_code = raw_data.get(id_field)
            source_display = raw_data.get(display_field)

            record: dict[str, Any] = {
                "entity_type": sheet_name,
                "source_code": self._clean_value(source_code),
                "source_display": self._clean_value(source_display),
                "raw_data": raw_data,
            }
            if section_label:
                record["section"] = section_label

            records.append(record)

        return records

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def extract_entity(self, sheet_name: str) -> list[dict[str, Any]]:
        """Extract records for a single entity / sheet.

        Parameters
        ----------
        sheet_name : str
            Name of the worksheet (must match one of the known entity
            types or a sheet present in the workbook).

        Returns
        -------
        list[dict]
            Each dict contains ``entity_type``, ``source_code``,
            ``source_display``, ``raw_data``, and optionally ``section``.
        """
        wb = self._open()

        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. "
                f"Available sheets: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # If we have a predefined config, use it
        if sheet_name in _SHEET_CONFIGS:
            all_records: list[dict[str, Any]] = []
            for section_cfg in _SHEET_CONFIGS[sheet_name]:
                section_records = self._read_section(ws, section_cfg, sheet_name)
                all_records.extend(section_records)
            logger.info(
                "Extracted %d records from sheet '%s'",
                len(all_records),
                sheet_name,
            )
            return all_records

        # Fallback: generic extraction for unknown sheets
        return self._extract_generic(ws, sheet_name)

    def _extract_generic(
        self, ws: Any, sheet_name: str
    ) -> list[dict[str, Any]]:
        """Fallback extractor for sheets without a predefined config.

        Heuristic: scan downward until a row is found where every cell
        looks like a header (non-numeric short strings).  Use that row as
        column headers, then read subsequent rows.
        """
        header_row: int | None = None
        headers: list[tuple[int, str]] = []  # (col_index, header_name)

        for row_idx in range(1, min(ws.max_row + 1, 20)):
            candidate: list[tuple[int, str]] = []
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    cleaned = str(val).strip()
                    if cleaned:
                        candidate.append((col_idx, cleaned))

            # Accept rows that have at least 2 header-like cells and none
            # look like pure integers (data rows typically start with IDs).
            if len(candidate) >= 2:
                looks_like_header = all(
                    not str(name).isdigit() for _, name in candidate
                )
                if looks_like_header:
                    header_row = row_idx
                    headers = candidate
                    break

        if header_row is None or not headers:
            logger.warning(
                "Could not detect header row for sheet '%s'; "
                "returning empty list.",
                sheet_name,
            )
            return []

        records: list[dict[str, Any]] = []
        id_field = headers[0][1]
        display_field = headers[-1][1] if len(headers) > 1 else headers[0][1]

        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_data: dict[str, Any] = {}
            all_none = True
            for col_idx, hdr_name in headers:
                cell_value = self._clean_value(
                    ws.cell(row=row_idx, column=col_idx).value
                )
                raw_data[hdr_name] = cell_value
                if cell_value is not None:
                    all_none = False

            if all_none:
                continue

            records.append(
                {
                    "entity_type": sheet_name,
                    "source_code": raw_data.get(id_field),
                    "source_display": raw_data.get(display_field),
                    "raw_data": raw_data,
                }
            )

        logger.info(
            "Extracted %d records from sheet '%s' (generic parser)",
            len(records),
            sheet_name,
        )
        return records

    def extract_all(self) -> dict[str, list[dict[str, Any]]]:
        """Extract records from every sheet in the workbook.

        Returns
        -------
        dict[str, list[dict]]
            Keys are sheet names; values are lists of record dicts.
        """
        wb = self._open()
        result: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in wb.sheetnames:
            try:
                result[sheet_name] = self.extract_entity(sheet_name)
            except Exception:
                logger.exception(
                    "Failed to extract sheet '%s'; skipping.", sheet_name
                )
                result[sheet_name] = []
        return result

    def profile(self) -> dict[str, dict[str, Any]]:
        """Return data profiling information for every entity sheet.

        Returns
        -------
        dict[str, dict]
            Keyed by sheet name.  Each value contains:
            - ``entity_type``
            - ``total_records``
            - ``non_null_codes`` -- count of records with a non-null
              ``source_code``
            - ``null_code_rate`` -- fraction of records with null codes
            - ``mapping_status`` -- ``"unmapped"`` / ``"partially_mapped"``
              / ``"mapped"`` based on whether the sheet's raw_data
              contains any recognised target-terminology columns with
              non-null values.
        """
        all_data = self.extract_all()
        profiles: dict[str, dict[str, Any]] = {}

        # Column names that indicate a target mapping already exists
        mapped_indicators = {
            "SnomedCode", "SNOMED", "FHIR_CODE", "FHIR_Code", "FHIR CODE",
            "Code", "code", "CODE", "LOINC", "LOINC_Code_Type",
            "LOINC_Code_Category", "AMTcode", "AMTMPPCODE",
            "AIR_CODE", "AIR CODE", "ACIRCODE", "System",
        }

        for sheet_name, records in all_data.items():
            total = len(records)
            non_null_codes = sum(
                1 for r in records if r.get("source_code") is not None
            )
            null_code_rate = (
                round(1 - non_null_codes / total, 4) if total > 0 else 0.0
            )

            # Determine mapping status by checking if any target-terminology
            # column already has non-null values
            mapped_count = 0
            for rec in records:
                raw = rec.get("raw_data", {})
                has_mapping = any(
                    raw.get(k) is not None
                    for k in mapped_indicators
                    if k in raw
                )
                if has_mapping:
                    mapped_count += 1

            if total == 0:
                status = "unmapped"
            elif mapped_count == 0:
                status = "unmapped"
            elif mapped_count < total:
                status = "partially_mapped"
            else:
                status = "mapped"

            profiles[sheet_name] = {
                "entity_type": sheet_name,
                "total_records": total,
                "non_null_codes": non_null_codes,
                "null_code_rate": null_code_rate,
                "mapping_status": status,
            }

        return profiles
